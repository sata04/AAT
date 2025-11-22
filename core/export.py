#!/usr/bin/env python3
"""
データエクスポートモジュール

処理されたデータと解析結果をExcelファイルにエクスポートする機能を提供します。
重力レベルデータとグラフ、G-quality解析結果を保存します。
"""

from __future__ import annotations

import shutil
from collections.abc import Callable
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from core.logger import get_logger, log_exception
from core.paths import ensure_graphs_dir, ensure_results_dir

# モジュール用のロガーを初期化
logger = get_logger("export")

ConfirmHandler = Callable[[Path], bool]
NotifyHandler = Callable[[str], None]


def _default_confirm_overwrite(path: Path) -> bool:
    logger.warning("上書き確認のハンドラが指定されていないため自動的に上書きします: %s", path)
    return True


def _default_notify_warning(message: str) -> None:
    logger.warning(message)


def _default_notify_info(message: str) -> None:
    logger.info(message)


def create_output_directories(csv_dir: str | None = None) -> tuple[Path, Path]:
    """
    出力用ディレクトリ構造を作成する

    CSVファイルがあるディレクトリに `results_AAT/graphs` ディレクトリを作成します。
    csv_dir が指定されていない場合は、プロジェクトルートにフォールバックします。

    Args:
        csv_dir (str, optional): CSVファイルのディレクトリパス。指定されていればそのディレクトリに作成

    Returns:
        tuple: 作成した結果ディレクトリとグラフディレクトリのパス
    """
    logger.debug("create_output_directories called with csv_dir: %s", csv_dir)

    results_dir = ensure_results_dir(csv_dir)
    graphs_dir = ensure_graphs_dir(csv_dir)

    logger.debug("Created directories: results=%s, graphs=%s", results_dir, graphs_dir)

    return results_dir, graphs_dir


def export_data(
    time: pd.Series,
    adjusted_time: pd.Series,
    gravity_level_inner_capsule: pd.Series,
    gravity_level_drag_shield: pd.Series,
    file_path: str,
    min_mean_inner_capsule: float | None,
    min_time_inner_capsule: float | None,
    min_std_inner_capsule: float | None,
    min_mean_drag_shield: float | None,
    min_time_drag_shield: float | None,
    min_std_drag_shield: float | None,
    graph_path: str | None,
    filtered_time: pd.Series,  # フィルタリング済みの時間データを追加  # noqa: ARG001
    filtered_adjusted_time: pd.Series,  # フィルタリング済みの調整時間データを追加  # noqa: ARG001
    config: dict[str, Any] | None = None,  # 設定パラメータを追加
    raw_data: pd.DataFrame | None = None,  # 元のCSVデータを追加
    *,
    confirm_overwrite: ConfirmHandler | None = None,
    notify_warning: NotifyHandler | None = None,
    notify_info: NotifyHandler | None = None,
) -> str:
    """
    処理されたデータとグラフをExcelにエクスポートする

    重力レベルデータと統計情報を含むExcelファイルを作成します。
    グラフは別途グラフディレクトリに保存します。
    既存のファイルが存在する場合は、上書きするかどうかを確認します。

    Args:
        time (pandas.Series): 時間データ
        adjusted_time (pandas.Series): 調整された時間データ
        gravity_level_inner_capsule (pandas.Series): Inner Capsuleの重力レベル
        gravity_level_drag_shield (pandas.Series): Drag Shieldの重力レベル
        file_path (str): 元のCSVファイルのパス
        min_mean_inner_capsule (float): Inner Capsuleの最小標準偏差ウィンドウの平均値
        min_time_inner_capsule (float): Inner Capsuleの最小標準偏差ウィンドウの開始時間
        min_std_inner_capsule (float): Inner Capsuleの最小標準偏差値
        min_mean_drag_shield (float): Drag Shieldの最小標準偏差ウィンドウの平均値
        min_time_drag_shield (float): Drag Shieldの最小標準偏差ウィンドウの開始時間
        min_std_drag_shield (float): Drag Shieldの最小標準偏差値
        graph_path (str): 保存されたグラフの画像ファイルパス
        filtered_time (pandas.Series): フィルタリングされた時間データ
        filtered_adjusted_time (pandas.Series): フィルタリングされた調整時間データ
        config (dict, optional): 設定パラメータ。指定されない場合はデフォルト値を使用
        raw_data (pandas.DataFrame, optional): 元のCSVデータ。加速度データ出力用

    Returns:
        str: 出力されたExcelファイルのパス

    Raises:
        ValueError: データのエクスポート中にエラーが発生した場合
    """
    confirm_overwrite = confirm_overwrite or _default_confirm_overwrite
    notify_warning = notify_warning or _default_notify_warning
    notify_info = notify_info or _default_notify_info

    # CSVファイルのディレクトリとファイル名を取得
    file_path_obj = Path(file_path)
    csv_dir = str(file_path_obj.parent)
    base_name = file_path_obj.stem

    # 出力ディレクトリ構造を作成
    results_dir, graphs_dir = create_output_directories(csv_dir)

    # 出力ファイルパスの設定（シンプルな名前を使用）
    output_file_path = results_dir / f"{base_name}.xlsx"

    # グラフファイルの新しいパスを設定（短い名前を使用）
    new_graph_path = graphs_dir / f"{base_name}_gl.png"

    # 既存グラフが元のパスに存在する場合、新しいパスにコピー
    if graph_path is not None:
        graph_path_obj = Path(graph_path)
        if graph_path_obj.exists() and graph_path_obj != new_graph_path:
            shutil.copy2(graph_path, new_graph_path)

    # 既存ファイルの確認
    if output_file_path.exists():
        if not confirm_overwrite(output_file_path):
            # 新しいファイル名を生成（連番を付加）
            counter = 1
            while (results_dir / f"{base_name}_{counter}.xlsx").exists():
                counter += 1
            output_file_path = results_dir / f"{base_name}_{counter}.xlsx"

    try:
        # 共通の時間軸を作成
        time_ranges = []
        if time is not None and not time.empty:
            time_ranges.append((time.min(), time.max()))
        if adjusted_time is not None and not adjusted_time.empty:
            time_ranges.append((adjusted_time.min(), adjusted_time.max()))

        if not time_ranges:
            raise ValueError("エクスポート可能な時間データがありません。")

        start_time = max(range_[0] for range_ in time_ranges)
        end_time = min(range_[1] for range_ in time_ranges)

        # オーバーラップがない場合は全体範囲を使用
        if end_time < start_time:
            start_time = min(range_[0] for range_ in time_ranges)
            end_time = max(range_[1] for range_ in time_ranges)

        # configが指定されていない場合のデフォルト値
        if config is None:
            config = {}

        # 時間間隔を計算（サンプリングレートに基づく）
        sampling_rate = config.get("sampling_rate", 1000)  # 設定からサンプリングレートを取得、デフォルトは1000Hz
        time_step = 1.0 / sampling_rate

        # 共通の時間軸を生成
        unified_time = np.arange(start_time, end_time + time_step, time_step)

        # データフレームの作成（統一された時間軸）
        export_columns = {"Time (s)": unified_time}
        if (
            time is not None
            and not time.empty
            and gravity_level_inner_capsule is not None
            and not gravity_level_inner_capsule.empty
        ):
            export_columns["Gravity Level (Inner Capsule) (G)"] = np.interp(
                unified_time, time, gravity_level_inner_capsule
            )
        if (
            adjusted_time is not None
            and not adjusted_time.empty
            and gravity_level_drag_shield is not None
            and not gravity_level_drag_shield.empty
        ):
            export_columns["Gravity Level (Drag Shield) (G)"] = np.interp(
                unified_time, adjusted_time, gravity_level_drag_shield
            )
        export_data = pd.DataFrame(export_columns)

        # 統計情報のデータフレームを作成
        stats_df = pd.DataFrame(
            {
                "Statistic": [
                    "Inner Capsule: Mean Gravity Level of the interval with the smallest standard deviation(G)",
                    "Inner Capsule: Time at smallest Standard Deviation(s)",
                    "Inner Capsule: smallest Standard Deviation(G)",
                    "Drag Shield: Mean Gravity Level of the interval with the smallest standard deviation(G)",
                    "Drag Shield: Time at smallest Standard Deviation(s)",
                    "Drag Shield: smallest Standard Deviation(G)",
                ],
                "Value": [
                    min_mean_inner_capsule,
                    min_time_inner_capsule,
                    min_std_inner_capsule,
                    min_mean_drag_shield,
                    min_time_drag_shield,
                    min_std_drag_shield,
                ],
            }
        )

        # トリミング範囲の加速度データを準備
        acceleration_data = None
        if raw_data is not None:
            # 元のCSVデータが提供されている場合
            try:
                # 設定から列名を取得
                time_column = config.get("time_column")
                acceleration_inner_column = config.get("acceleration_column_inner_capsule")
                acceleration_drag_column = config.get("acceleration_column_drag_shield")
                use_inner = config.get("use_inner_acceleration", True)
                use_drag = config.get("use_drag_acceleration", True)

                logger.info(
                    "加速度データの処理開始: "
                    f"時間列={time_column}, 内カプセル加速度列={acceleration_inner_column}, 外カプセル加速度列={acceleration_drag_column}, "
                    f"_inner使用={use_inner}, _drag使用={use_drag}"
                )
                logger.debug(f"元データの列: {raw_data.columns.tolist()}")

                if time_column is None:
                    notify_warning("加速度データの時間列が設定されていないため、エクスポートをスキップします。")
                    raw_columns_valid = False
                else:
                    raw_columns_valid = True
                    missing_cols = []
                    if time_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"時間列({time_column})")
                    if use_inner and acceleration_inner_column and acceleration_inner_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"内カプセル加速度列({acceleration_inner_column})")
                    if use_drag and acceleration_drag_column and acceleration_drag_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"外カプセル加速度列({acceleration_drag_column})")

                    if not raw_columns_valid and missing_cols:
                        logger.warning(f"必要な列が見つかりません: {missing_cols}")
                        notify_warning(
                            "必要な列が見つかりません: "
                            f"\n{', '.join(missing_cols)}\n\n"
                            "加速度データがシートに追加されません。\n"
                            "CSVファイルを選択する際、正しい列を選んでください。"
                        )

                if raw_columns_valid:
                    try:
                        orig_time_data = raw_data[time_column].values.astype(float)
                        acceleration_columns: dict[str, np.ndarray] = {}

                        if use_inner and acceleration_inner_column:
                            acceleration_columns["inner"] = raw_data[acceleration_inner_column].values.astype(float)
                        if use_drag and acceleration_drag_column:
                            acceleration_columns["drag"] = raw_data[acceleration_drag_column].values.astype(float)

                        if not acceleration_columns:
                            logger.info(
                                "有効な加速度列が選択されていないため、加速度データのエクスポートをスキップします"
                            )
                        else:
                            orig_adjusted_time = orig_time_data
                            if "drag" in acceleration_columns:
                                acc_thresh = config.get("acceleration_threshold", 1.0)
                                sync_mask = np.abs(acceleration_columns["drag"]) < acc_thresh
                                if sync_mask.any():
                                    sync_idx = np.where(sync_mask)[0][0]
                                    orig_adjusted_time = orig_time_data - orig_time_data[sync_idx]
                                else:
                                    orig_adjusted_time = orig_time_data - orig_time_data[0]

                            accel_frame = {"Time (s)": unified_time}
                            if "inner" in acceleration_columns:
                                accel_frame["Acceleration (Inner Capsule) (m/s²)"] = np.interp(
                                    unified_time, orig_time_data, acceleration_columns["inner"]
                                )
                            if "drag" in acceleration_columns:
                                accel_frame["Acceleration (Drag Shield) (m/s²)"] = np.interp(
                                    unified_time, orig_adjusted_time, acceleration_columns["drag"]
                                )

                            acceleration_data = pd.DataFrame(accel_frame)
                            logger.info(f"共通時間軸で加速度データを作成: {len(acceleration_data)}行")

                    except Exception as e:
                        log_exception(e, "加速度データのエクスポート中にエラーが発生しました")
                        notify_warning(f"加速度データの保存中にエラーが発生しました: {e}")
                        acceleration_data = None
            except Exception as e:
                log_exception(e, "加速度データの準備中にエラーが発生しました")
                acceleration_data = None

        # Excelファイルにデータと統計情報を書き込む
        with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
            export_data.to_excel(writer, sheet_name="Gravity Level Data", index=False)
            stats_df.to_excel(writer, sheet_name="Gravity Level Statistics", index=False)
            if acceleration_data is not None:
                acceleration_data.to_excel(writer, sheet_name="Acceleration Data", index=False)
                logger.info(f"加速度データをシートに追加しました: {len(acceleration_data)}行")
            else:
                logger.warning("加速度データが作成されなかったため、シートに追加されません")

        graph_exists = graph_path is not None and Path(graph_path).exists()
        graph_display_target = new_graph_path if graph_exists else new_graph_path.parent
        graphs_message = (
            f"- グラフ画像: {graph_display_target}" if graph_exists else f"- グラフ出力フォルダ: {graph_display_target}"
        )
        message = f"保存が完了しました。\n- Gravity Levelデータ: {output_file_path}\n{graphs_message}"
        notify_info(message)

        return str(output_file_path)
    except PermissionError as e:
        error_msg = f"{output_file_path} に書き込みできません。権限を確認してください。"
        logger.error(error_msg)
        raise ValueError(error_msg) from e
    except Exception as e:
        error_msg = f"データの保存中にエラーが発生しました: {e}"
        log_exception(e, error_msg)
        raise ValueError(error_msg) from e


def export_g_quality_data(g_quality_data, original_file_path, g_quality_graph_path=None):
    """
    G-quality解析の結果をエクスポートする

    異なるウィンドウサイズでのG-quality評価結果をExcelファイルに追加または新規作成します。
    既存のExcelファイルがある場合は、G-quality Analysis シートを更新します。
    また、G-qualityグラフが提供されている場合は、指定されたディレクトリにコピーします。

    Args:
        g_quality_data (list): G-quality解析の結果データ
            各要素は (window_size, min_time_inner_capsule, min_mean_inner_capsule, min_std_inner_capsule,
                      min_time_drag_shield, min_mean_drag_shield, min_std_drag_shield) の形式のタプル
        original_file_path (str): 元のCSVファイルのパス
        g_quality_graph_path (str, optional): G-qualityグラフの画像ファイルパス。指定された場合はグラフをコピーします。

    Returns:
        str or None: 出力されたExcelファイルのパス、または失敗した場合はNone

    Raises:
        ValueError: データのエクスポート中にエラーが発生した場合
    """
    # CSVファイルのディレクトリとファイル名を取得
    file_path_obj = Path(original_file_path)
    csv_dir = str(file_path_obj.parent)
    base_name = file_path_obj.stem

    # 出力ディレクトリ構造を作成
    results_dir, graphs_dir = create_output_directories(csv_dir)

    # G-qualityグラフの処理
    if g_quality_graph_path and Path(g_quality_graph_path).exists():
        # グラフファイルの新しいパスを設定（短い名前を使用）
        new_graph_path = graphs_dir / f"{base_name}_gq.png"

        # 同じファイルでない場合のみコピーを実行
        import shutil

        try:
            # パスを正規化して比較
            source_path = Path(g_quality_graph_path).resolve()
            dest_path = new_graph_path.resolve()

            if source_path != dest_path:
                shutil.copy2(g_quality_graph_path, new_graph_path)
                logger.info(f"G-qualityグラフを保存しました: {g_quality_graph_path} -> {new_graph_path}")
            else:
                logger.debug(f"G-qualityグラフは既に正しい場所にあります: {new_graph_path}")
        except Exception as e:
            logger.warning(f"G-qualityグラフの保存中にエラーが発生しました: {e}")

    # データフレームの作成
    df = pd.DataFrame(
        g_quality_data,
        columns=[
            "Window Size (s)",
            "Inner Capsule: Time at smallest Standard Deviation(s)",
            "Inner Capsule: Mean Gravity Level of the interval with the smallest standard deviation(G)",
            "Inner Capsule: smallest Standard Deviation(G)",
            "Drag Shield: Time at smallest Standard Deviation(s)",
            "Drag Shield: Mean Gravity Level of the interval with the smallest standard deviation(G)",
            "Drag Shield: smallest Standard Deviation(G)",
        ],
    )

    # 出力ファイルパスの設定
    output_file_path = results_dir / f"{base_name}.xlsx"

    try:
        # 既存のExcelファイルを読み込むか、新規作成
        try:
            workbook = load_workbook(output_file_path)
        except FileNotFoundError:
            workbook = Workbook()
            # デフォルトのシートを削除（後で必要なシートを追加する）
            sheet_to_remove = workbook.active
            if sheet_to_remove is not None:
                workbook.remove(sheet_to_remove)

        # G-quality Analysis シートを作成または更新
        sheet_name = "G-quality Analysis"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 既存のシートをクリア
            for row in sheet[sheet.min_row : sheet.max_row]:
                for cell in row:
                    cell.value = None
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        # データをシートに書き込む（1行目から開始）
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

        # ファイルを保存
        workbook.save(output_file_path)
        return output_file_path
    except PermissionError as e:
        raise ValueError(f"{output_file_path} に書き込みできません。ファイルが開かれている可能性があります。") from e
    except Exception as e:
        raise ValueError(f"データの保存中にエラーが発生しました: {e}") from e
