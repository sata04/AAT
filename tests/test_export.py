from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.export import create_output_directories, export_data, export_g_quality_data


def test_create_output_directories_respects_csv_dir(tmp_path):
    csv_dir = tmp_path / "csvs"
    csv_dir.mkdir()

    results_dir, graphs_dir = create_output_directories(str(csv_dir))

    assert results_dir == csv_dir / "results_AAT"
    assert graphs_dir == results_dir / "graphs"
    assert results_dir.exists()
    assert graphs_dir.exists()


def test_export_data_writes_excel_with_sheets(sample_config, raw_data_frame, tmp_path, dummy_message_box):
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    time_series = raw_data_frame["time_s"]
    adjusted_time = time_series - time_series.iloc[0]
    gl_ic = raw_data_frame["acc_ic"] / sample_config["gravity_constant"]
    gl_ds = raw_data_frame["acc_ds"] / sample_config["gravity_constant"]

    output_path = export_data(
        time=time_series,
        adjusted_time=adjusted_time,
        gravity_level_inner_capsule=gl_ic,
        gravity_level_drag_shield=gl_ds,
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=time_series,
        filtered_adjusted_time=adjusted_time,
        config=sample_config,
        raw_data=raw_data_frame,
    )

    assert Path(output_path).exists()
    workbook = load_workbook(output_path)
    assert "Gravity Level Data" in workbook.sheetnames
    assert "Gravity Level Statistics" in workbook.sheetnames
    assert "Acceleration Data" in workbook.sheetnames


def test_export_copies_graph_to_results_dir(sample_config, raw_data_frame, tmp_path, dummy_message_box):
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    graph_path = tmp_path / "graph.png"
    graph_path.write_bytes(b"pngdata")

    time_series = raw_data_frame["time_s"]
    adjusted_time = time_series - time_series.iloc[0]
    gl_ic = raw_data_frame["acc_ic"] / sample_config["gravity_constant"]
    gl_ds = raw_data_frame["acc_ds"] / sample_config["gravity_constant"]

    output_path = export_data(
        time=time_series,
        adjusted_time=adjusted_time,
        gravity_level_inner_capsule=gl_ic,
        gravity_level_drag_shield=gl_ds,
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=str(graph_path),
        filtered_time=time_series,
        filtered_adjusted_time=adjusted_time,
        config=sample_config,
        raw_data=raw_data_frame,
    )

    graphs_dir = Path(output_path).parent / "graphs"
    copied_graph = graphs_dir / f"{csv_path.stem}_gl.png"
    assert copied_graph.exists()


def test_export_skips_acceleration_sheet_when_columns_missing(sample_config, tmp_path, dummy_message_box):
    csv_path = tmp_path / "data.csv"
    data = pd.DataFrame({"time_s": [0.0, 0.1], "acc_x": [1.0, 1.1]})
    data.to_csv(csv_path, index=False)

    time_series = data["time_s"]
    adjusted_time = time_series - time_series.iloc[0]
    gl_ic = pd.Series([0.0, 0.1])
    gl_ds = pd.Series([0.0, 0.1])

    output_path = export_data(
        time=time_series,
        adjusted_time=adjusted_time,
        gravity_level_inner_capsule=gl_ic,
        gravity_level_drag_shield=gl_ds,
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=time_series,
        filtered_adjusted_time=adjusted_time,
        config=sample_config,
        raw_data=data,
    )

    workbook = load_workbook(output_path)
    assert "Gravity Level Data" in workbook.sheetnames
    assert "Acceleration Data" not in workbook.sheetnames


def test_export_g_quality_data_creates_new_file(tmp_path):
    """Test exporting G-quality data to a new Excel file."""
    original_csv = tmp_path / "test.csv"
    original_csv.touch()

    g_quality_data = [
        (0.1, 0.0, 0.05, 0.001, 0.0, 0.06, 0.003),  # window, time_ic, mean_ic, std_ic, time_ds, mean_ds, std_ds
        (0.2, 0.1, 0.04, 0.001, 0.1, 0.05, 0.003),
    ]

    output_path = export_g_quality_data(g_quality_data, str(original_csv))

    assert Path(output_path).exists()
    workbook = load_workbook(output_path)
    assert "G-quality Analysis" in workbook.sheetnames

    sheet = workbook["G-quality Analysis"]
    # Header + 2 data rows
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=1).value == 0.1  # Window size
    assert sheet.cell(row=2, column=3).value == 0.05  # Mean IC


def test_export_g_quality_data_appends_to_existing_file(tmp_path, sample_config, raw_data_frame):
    """Test appending G-quality data to an existing Excel file."""
    # First create a file with standard export
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    time_series = raw_data_frame["time_s"]
    gl_ic = raw_data_frame["acc_ic"]
    gl_ds = raw_data_frame["acc_ds"]

    # Create initial export
    output_path = export_data(
        time=time_series,
        adjusted_time=time_series,
        gravity_level_inner_capsule=gl_ic,
        gravity_level_drag_shield=gl_ds,
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=time_series,
        filtered_adjusted_time=time_series,
        config=sample_config,
    )

    # Now append G-quality data
    g_quality_data = [(0.1, 0.0, 0.05, 0.001, 0.0, 0.06, 0.003)]
    updated_path = export_g_quality_data(g_quality_data, str(csv_path))

    assert str(updated_path) == str(output_path)
    workbook = load_workbook(updated_path)
    assert "Gravity Level Data" in workbook.sheetnames
    assert "G-quality Analysis" in workbook.sheetnames


def test_export_data_handles_write_permission_error(tmp_path, sample_config, raw_data_frame, monkeypatch):
    """Test handling of permission error during export."""
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    # Mock pd.ExcelWriter to raise PermissionError on init
    def mock_excel_writer(*args, **kwargs):
        raise PermissionError("Permission denied")

    monkeypatch.setattr("pandas.ExcelWriter", mock_excel_writer)

    with pytest.raises(ValueError) as exc_info:
        export_data(
            time=raw_data_frame["time_s"],
            adjusted_time=raw_data_frame["time_s"],
            gravity_level_inner_capsule=raw_data_frame["acc_ic"],
            gravity_level_drag_shield=raw_data_frame["acc_ds"],
            file_path=str(csv_path),
            min_mean_inner_capsule=0.1,
            min_time_inner_capsule=0.0,
            min_std_inner_capsule=0.01,
            min_mean_drag_shield=0.2,
            min_time_drag_shield=0.0,
            min_std_drag_shield=0.02,
            graph_path=None,
            filtered_time=raw_data_frame["time_s"],
            filtered_adjusted_time=raw_data_frame["time_s"],
            config=sample_config,
        )

    assert "書き込みできません" in str(exc_info.value)


def test_export_data_respects_overwrite_confirmation(tmp_path, sample_config, raw_data_frame):
    """Test that export respects the overwrite confirmation callback."""
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    # Create the target excel file first
    results_dir = csv_path.parent / "results_AAT"
    results_dir.mkdir(parents=True)
    excel_path = results_dir / "data.xlsx"
    excel_path.touch()

    # Callback that denies overwrite
    confirm_called = False

    def deny_overwrite(path):
        nonlocal confirm_called
        confirm_called = True
        return False

    # Should return the existing path without modifying it
    result_path = export_data(
        time=raw_data_frame["time_s"],
        adjusted_time=raw_data_frame["time_s"],
        gravity_level_inner_capsule=raw_data_frame["acc_ic"],
        gravity_level_drag_shield=raw_data_frame["acc_ds"],
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=raw_data_frame["time_s"],
        filtered_adjusted_time=raw_data_frame["time_s"],
        config=sample_config,
        confirm_overwrite=deny_overwrite,
    )

    assert confirm_called
    # When overwrite is denied, it creates a new file with a counter
    assert result_path != str(excel_path)
    assert "_1.xlsx" in result_path


def test_export_data_handles_missing_graph_file(tmp_path, sample_config, raw_data_frame):
    """Test that export proceeds even if graph file is missing."""
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    non_existent_graph = tmp_path / "non_existent.png"

    notify_called = False

    def notify_warning(msg):
        nonlocal notify_called
        notify_called = True
        # Warning might not be called for missing graph, but info should be

    export_data(
        time=raw_data_frame["time_s"],
        adjusted_time=raw_data_frame["time_s"],
        gravity_level_inner_capsule=raw_data_frame["acc_ic"],
        gravity_level_drag_shield=raw_data_frame["acc_ds"],
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=str(non_existent_graph),
        filtered_time=raw_data_frame["time_s"],
        filtered_adjusted_time=raw_data_frame["time_s"],
        config=sample_config,
        notify_warning=notify_warning,
    )

    # Function should complete without error
    results_dir = csv_path.parent / "results_AAT"
    assert (results_dir / "data.xlsx").exists()


def test_default_callbacks(tmp_path):
    """Test default callback functions."""
    from core.export import _default_confirm_overwrite, _default_notify_info, _default_notify_warning

    # These should just log and not raise errors
    assert _default_confirm_overwrite(tmp_path) is True
    _default_notify_warning("test warning")
    _default_notify_info("test info")


def test_export_data_without_config(tmp_path, raw_data_frame):
    """Test export_data without config argument."""
    csv_path = tmp_path / "data.csv"
    raw_data_frame.to_csv(csv_path, index=False)

    output_path = export_data(
        time=raw_data_frame["time_s"],
        adjusted_time=raw_data_frame["time_s"],
        gravity_level_inner_capsule=raw_data_frame["acc_ic"],
        gravity_level_drag_shield=raw_data_frame["acc_ds"],
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=raw_data_frame["time_s"],
        filtered_adjusted_time=raw_data_frame["time_s"],
    )

    assert Path(output_path).exists()


def test_export_data_with_no_overlap(tmp_path, sample_config):
    """Test export_data when time ranges do not overlap."""
    csv_path = tmp_path / "data.csv"

    # Create non-overlapping time series
    time1 = pd.Series([0.0, 1.0])
    time2 = pd.Series([10.0, 11.0])

    output_path = export_data(
        time=time1,
        adjusted_time=time2,
        gravity_level_inner_capsule=pd.Series([1.0, 1.0]),
        gravity_level_drag_shield=pd.Series([1.0, 1.0]),
        file_path=str(csv_path),
        min_mean_inner_capsule=0.1,
        min_time_inner_capsule=0.0,
        min_std_inner_capsule=0.01,
        min_mean_drag_shield=0.2,
        min_time_drag_shield=0.0,
        min_std_drag_shield=0.02,
        graph_path=None,
        filtered_time=time1,
        filtered_adjusted_time=time2,
        config=sample_config,
    )

    assert Path(output_path).exists()
    workbook = load_workbook(output_path)
    sheet = workbook["Gravity Level Data"]
    # Should cover the full range from min(time1) to max(time2)
    # 0.0 to 11.0 -> 12 rows + header = 13 rows (approx, depending on sampling rate)
    assert sheet.max_row > 10
